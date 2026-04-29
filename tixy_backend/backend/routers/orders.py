from datetime import date, datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import func, cast, Date
from sqlalchemy.orm import Session, joinedload

from core.database import get_db
from core.deps import get_current_user, require_manager, require_vendor
from models.order import Order, OrderLine, OrderStatus
from models.reference import Reference
from models.user import User, UserRole
from schemas.order import OrderCreate, OrderOut, OrderSummary, OrderUpdate

from pydantic import BaseModel, EmailStr

class SendToClientPayload(BaseModel):
    email: EmailStr

router = APIRouter(prefix="/orders", tags=["orders"])


# ── helpers ──────────────────────────────────────────────────────────────────

from models.client import Store, Client


def _load_order(db: Session, order_id: int) -> Order:
    order = (
        db.query(Order)
        .options(
            joinedload(Order.lines).joinedload(OrderLine.reference),
            joinedload(Order.vendor),
            joinedload(Order.store).joinedload(Store.client),
        )
        .filter(Order.id == order_id)
        .first()
    )
    if not order:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")
    return order


def _next_order_number(db: Session) -> str:
    last = db.query(func.max(Order.id)).scalar() or 752
    return str(last + 1).zfill(4)


# ── CRUD ─────────────────────────────────────────────────────────────────────

@router.get("/", response_model=list[OrderSummary])
def list_orders(
    vendor_id:     Optional[int]         = None,
    collection_id: Optional[int]         = None,
    status:        Optional[OrderStatus] = None,
    city:          Optional[str]         = None,
    date_from:     Optional[date]        = None,
    date_to:       Optional[date]        = None,
    db:  Session = Depends(get_db),
    me:  User    = Depends(get_current_user),
):
    q = db.query(Order).options(
        joinedload(Order.vendor),
        joinedload(Order.store).joinedload(Store.client),
        joinedload(Order.lines),
    )

    # vendedor solo ve sus propios pedidos
    if me.role == UserRole.VENDOR:
        q = q.filter(Order.vendor_id == me.id)
    elif vendor_id:
        q = q.filter(Order.vendor_id == vendor_id)

    if collection_id:
        q = q.filter(Order.collection_id == collection_id)
    if status:
        q = q.filter(Order.status == status)
    if city:
        store_ids = db.query(Store.id).filter(Store.city.ilike(f"%{city}%")).subquery()
        q = q.filter(Order.store_id.in_(store_ids))
    if date_from:
        q = q.filter(cast(Order.created_at, Date) >= date_from)
    if date_to:
        q = q.filter(cast(Order.created_at, Date) <= date_to)

    return q.order_by(Order.created_at.desc()).all()


@router.get("/{order_id}", response_model=OrderOut)
def get_order(
    order_id: int,
    db: Session = Depends(get_db),
    me: User    = Depends(get_current_user),
):
    order = _load_order(db, order_id)
    if me.role == UserRole.VENDOR and order.vendor_id != me.id:
        raise HTTPException(status_code=403, detail="Acceso denegado")
    return order


@router.post("/", response_model=OrderOut, status_code=201)
def create_order(
    payload: OrderCreate,
    db:  Session = Depends(get_db),
    me:  User    = Depends(require_vendor),
):
    order = Order(
        order_number=_next_order_number(db),
        vendor_id=me.id,
        store_id=payload.store_id,
        collection_id=payload.collection_id,
        notes=payload.notes,
        status=OrderStatus.DRAFT,
    )
    db.add(order)
    db.flush()

    for ln in payload.lines:
        ref = db.get(Reference, ln.reference_id)
        if not ref:
            raise HTTPException(status_code=400, detail=f"Referencia {ln.reference_id} no existe")
        db.add(OrderLine(
            order_id=order.id,
            reference_id=ln.reference_id,
            quantity=ln.quantity,
            unit_price=ln.unit_price,
        ))

    db.commit()
    return _load_order(db, order.id)


@router.patch("/{order_id}", response_model=OrderOut)
def update_order(
    order_id: int,
    payload:  OrderUpdate,
    db: Session = Depends(get_db),
    me: User    = Depends(require_vendor),
):
    """Vendedor edita un pedido propio en estado DRAFT o SENT.
    Si estaba SENT, regresa a DRAFT automáticamente para que lo re-envíe."""
    order = _load_order(db, order_id)
    if order.vendor_id != me.id:
        raise HTTPException(status_code=403, detail="Acceso denegado")
    if order.status == OrderStatus.CANCELLED:
        raise HTTPException(
            status_code=400,
            detail="No se puede editar un pedido cancelado",
        )

    if payload.store_id is not None:
        order.store_id = payload.store_id
    if payload.collection_id is not None:
        order.collection_id = payload.collection_id
    if payload.notes is not None:
        order.notes = payload.notes

    if payload.lines is not None:
        # Reemplazar todas las líneas existentes
        for line in list(order.lines):
            db.delete(line)
        db.flush()
        for ln in payload.lines:
            ref = db.get(Reference, ln.reference_id)
            if not ref:
                raise HTTPException(
                    status_code=400,
                    detail=f"Referencia {ln.reference_id} no existe",
                )
            db.add(OrderLine(
                order_id=order.id,
                reference_id=ln.reference_id,
                quantity=ln.quantity,
                unit_price=ln.unit_price,
            ))

    # Si estaba enviado y se edita, vuelve a borrador
    if order.status == OrderStatus.SENT:
        order.status  = OrderStatus.DRAFT
        order.sent_at = None

    db.commit()
    return _load_order(db, order_id)


@router.post("/{order_id}/send", response_model=OrderOut)
def send_order(
    order_id: int,
    db: Session = Depends(get_db),
    me: User    = Depends(require_vendor),
):
    """Vendedor marca el pedido como enviado."""
    order = _load_order(db, order_id)
    if me.role == UserRole.VENDOR and order.vendor_id != me.id:
        raise HTTPException(status_code=403, detail="Acceso denegado")
    if order.status != OrderStatus.DRAFT:
        raise HTTPException(status_code=400, detail="Solo se pueden enviar pedidos en borrador")
    order.status  = OrderStatus.SENT
    order.sent_at = datetime.now(timezone.utc)
    db.commit()
    return _load_order(db, order_id)



@router.post("/{order_id}/cancel", response_model=OrderOut)
def cancel_order(
    order_id: int,
    db: Session = Depends(get_db),
    me: User    = Depends(get_current_user),
):
    order = _load_order(db, order_id)
    if me.role == UserRole.VENDOR and order.vendor_id != me.id:
        raise HTTPException(status_code=403, detail="Acceso denegado")
    if order.status == OrderStatus.CANCELLED:
        raise HTTPException(status_code=400, detail="El pedido ya está cancelado")
    order.status = OrderStatus.CANCELLED
    db.commit()
    return _load_order(db, order_id)


# ── Enviar PDF al cliente por email ──────────────────────────────────────────

@router.post("/{order_id}/send-to-client", status_code=200)
def send_order_to_client(
    order_id: int,
    payload:  SendToClientPayload,
    db: Session = Depends(get_db),
    me: User    = Depends(require_vendor),
):
    """Genera el PDF de la orden y lo envía por email al cliente."""
    from routers.pdf import _build_pdf
    from core.email import send_order_pdf_email
    from models.client import Store

    order = (
        db.query(Order)
        .options(
            joinedload(Order.lines).joinedload(OrderLine.reference),
            joinedload(Order.vendor),
            joinedload(Order.store).joinedload(Store.client),
        )
        .filter(Order.id == order_id)
        .first()
    )
    if not order:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")
    if me.role == UserRole.VENDOR and order.vendor_id != me.id:
        raise HTTPException(status_code=403, detail="Acceso denegado")

    try:
        pdf_bytes = _build_pdf(order, show_total=False)
        client_name = order.store.client.business_name if order.store and order.store.client else ""
        send_order_pdf_email(
            to_email=payload.email,
            client_name=client_name,
            order_number=order.order_number,
            pdf_bytes=pdf_bytes,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar el correo: {str(e)}")

    return {"ok": True, "detail": f"Orden #{order.order_number} enviada a {payload.email}"}


# ── Reporte Excel (gerencia) ─────────────────────────────────────────────────

@router.get("/report/excel", tags=["reports"])
def download_excel_report(
    collection_id: int,
    report_type:   str = Query("unidades", pattern="^(unidades|costos)$"),
    db: Session = Depends(get_db),
    _:  User    = Depends(require_manager),
):
    """
    Genera un Excel con el reporte de ventas por colección.
    - report_type='unidades': cantidades vendidas por referencia × cliente
    - report_type='costos':   valor en pesos por referencia × cliente
    Una pestaña por vendedor.
    """
    import io
    from collections import defaultdict
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from models.collection import Collection
    from models.reference import Reference

    # ── Paleta Tixy ───────────────────────────────────────────────────────
    PINK       = "C0206A"
    PINK_LITE  = "FAE0EE"
    DARK       = "1A0D14"
    WHITE      = "FFFFFF"
    GRAY_LIGHT = "F5F0F2"
    GRAY_MID   = "E8D8E0"

    thin = Side(style="thin", color="D0B8C8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfill(color): return PatternFill("solid", fgColor=color)
    def font(bold=False, color=DARK, size=10): return Font(bold=bold, color=color, size=size, name="Arial")
    def align(h="left", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ── Verificar colección ───────────────────────────────────────────────
    collection = db.get(Collection, collection_id)
    if not collection:
        raise HTTPException(status_code=404, detail="Colección no encontrada")

    # ── Consulta pivot ────────────────────────────────────────────────────
    from sqlalchemy import text as sa_text
    rows = (
        db.query(
            User.id.label("vendor_id"),
            User.full_name.label("vendor_name"),
            Reference.id.label("ref_id"),
            Reference.code.label("ref_code"),
            Reference.description.label("ref_desc"),
            Reference.base_price.label("ref_price"),
            Order.store_id,
            Store.name.label("store_name"),
            Client.business_name.label("client_name"),
            func.sum(OrderLine.quantity).label("total_units"),
            func.sum(OrderLine.quantity * OrderLine.unit_price).label("total_value"),
        )
        .join(Order, Order.vendor_id == User.id)
        .join(OrderLine, OrderLine.order_id == Order.id)
        .join(Reference, Reference.id == OrderLine.reference_id)
        .join(Store, Store.id == Order.store_id)
        .join(Client, Client.id == Store.client_id)
        .filter(
            Order.collection_id == collection_id,
            Order.status != OrderStatus.CANCELLED,
        )
        .group_by(User.id, Reference.id, Order.store_id)
        .order_by(User.full_name, Reference.code, Store.name)
        .all()
    )

    if not rows:
        raise HTTPException(status_code=404, detail="No hay datos para esta colección")

    # ── Estructurar datos por vendedor ────────────────────────────────────
    # vendors[vendor_id] = { name, refs: {ref_id: {code, desc, price}}, stores: {store_id: name}, data: {(ref_id,store_id): val} }
    vendors: dict = {}
    for r in rows:
        if r.vendor_id not in vendors:
            vendors[r.vendor_id] = {
                "name":   r.vendor_name,
                "refs":   {},      # ref_id -> (code, desc, price)
                "stores": {},      # store_id -> name
                "data":   defaultdict(float),  # (ref_id, store_id) -> valor
            }
        v = vendors[r.vendor_id]
        v["refs"][r.ref_id]   = (r.ref_code, r.ref_desc, float(r.ref_price or 0))
        v["stores"][r.store_id] = f"{r.client_name} / {r.store_name}"
        value = float(r.total_units) if report_type == "unidades" else float(r.total_value)
        v["data"][(r.ref_id, r.store_id)] = value

    # ── Construir workbook ────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja vacía inicial

    is_units = report_type == "unidades"
    num_fmt  = "#,##0" if is_units else "$#,##0"

    for vendor_id, vdata in vendors.items():
        # Nombre de pestaña (máx 31 chars, sin caracteres inválidos)
        safe_name = vdata["name"][:31].replace("/", "-").replace(":", "-").replace("[", "").replace("]", "")
        ws = wb.create_sheet(title=safe_name)

        refs   = sorted(vdata["refs"].items(),   key=lambda x: x[1][0])   # ordenar por código
        stores = sorted(vdata["stores"].items(), key=lambda x: x[1])       # ordenar por nombre

        n_stores = len(stores)
        # Columnas: PRECIO(A) | DESCRIPCIÓN(B) | REF(C) | store1..storeN | TOTAL
        COL_PRECIO = 1; COL_DESC = 2; COL_REF = 3
        COL_FIRST_STORE = 4
        COL_TOTAL = COL_FIRST_STORE + n_stores
        total_cols = COL_TOTAL

        # ── Fila 1: Título del reporte ──────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(1, 1)
        col_label = collection.name
        tipo_label = "Unidades Vendidas" if is_units else "Ventas en Costos"
        title_cell.value = f"TIXY GLAMOUR · {col_label.upper()} · {tipo_label.upper()}"
        title_cell.font      = Font(bold=True, color=WHITE, size=12, name="Arial")
        title_cell.fill      = hfill(DARK)
        title_cell.alignment = align("center", "center")
        ws.row_dimensions[1].height = 22

        # ── Fila 2: Sub-título vendedor ─────────────────────────────────
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        sub_cell = ws.cell(2, 1)
        sub_cell.value     = f"Vendedor: {vdata['name']}"
        sub_cell.font      = Font(bold=True, color=WHITE, size=10, name="Arial")
        sub_cell.fill      = hfill(PINK)
        sub_cell.alignment = align("center", "center")
        ws.row_dimensions[2].height = 18

        # ── Fila 3: Encabezados de columnas ─────────────────────────────
        HDR_ROW = 3
        headers = ["PRECIO BASE", "DESCRIPCIÓN", "REF"]
        headers += [s[1] for s in stores]
        headers += ["TOTAL"]

        for col_i, hdr in enumerate(headers, start=1):
            c = ws.cell(HDR_ROW, col_i, hdr)
            c.font      = font(bold=True, color=WHITE, size=9)
            c.fill      = hfill(PINK)
            c.alignment = align("center", "center", wrap=True)
            c.border    = border
        ws.row_dimensions[HDR_ROW].height = 40

        # ── Filas de datos ───────────────────────────────────────────────
        DATA_START = 4
        for row_i, (ref_id, (ref_code, ref_desc, ref_price)) in enumerate(refs, start=DATA_START):
            fill_row = hfill(GRAY_LIGHT) if row_i % 2 == 0 else None

            # PRECIO
            c = ws.cell(row_i, COL_PRECIO, ref_price)
            c.font = font(); c.alignment = align("right")
            c.number_format = "$#,##0"; c.border = border
            if fill_row: c.fill = fill_row

            # DESCRIPCIÓN
            c = ws.cell(row_i, COL_DESC, ref_desc)
            c.font = font(); c.alignment = align("left", wrap=True)
            c.border = border
            if fill_row: c.fill = fill_row

            # REF
            c = ws.cell(row_i, COL_REF, ref_code)
            c.font = Font(bold=True, color=PINK, size=10, name="Arial")
            c.alignment = align("center")
            c.border = border
            if fill_row: c.fill = fill_row

            # Valores por almacén
            store_cols = []
            for col_j, (store_id, _) in enumerate(stores, start=COL_FIRST_STORE):
                val = vdata["data"].get((ref_id, store_id), 0)
                c = ws.cell(row_i, col_j, val if val else None)
                c.font = font(); c.alignment = align("right")
                c.number_format = num_fmt; c.border = border
                if fill_row: c.fill = fill_row
                store_cols.append(get_column_letter(col_j))

            # TOTAL fila
            first_store_letter = get_column_letter(COL_FIRST_STORE)
            last_store_letter  = get_column_letter(COL_FIRST_STORE + n_stores - 1)
            total_formula = f"=SUM({first_store_letter}{row_i}:{last_store_letter}{row_i})"
            c = ws.cell(row_i, COL_TOTAL, total_formula)
            c.font = Font(bold=True, color=DARK, size=10, name="Arial")
            c.alignment = align("right")
            c.number_format = num_fmt; c.border = border
            c.fill = hfill(PINK_LITE)

        # ── Fila de totales ──────────────────────────────────────────────
        TOTAL_ROW = DATA_START + len(refs)
        ws.merge_cells(start_row=TOTAL_ROW, start_column=1, end_row=TOTAL_ROW, end_column=COL_REF)
        c = ws.cell(TOTAL_ROW, 1, "TOTALES")
        c.font = Font(bold=True, color=WHITE, size=10, name="Arial")
        c.fill = hfill(DARK); c.alignment = align("center", "center"); c.border = border

        for col_j in range(COL_FIRST_STORE, COL_TOTAL + 1):
            col_letter = get_column_letter(col_j)
            formula = f"=SUM({col_letter}{DATA_START}:{col_letter}{TOTAL_ROW - 1})"
            c = ws.cell(TOTAL_ROW, col_j, formula)
            c.font = Font(bold=True, color=WHITE, size=10, name="Arial")
            c.fill = hfill(DARK); c.alignment = align("right")
            c.number_format = num_fmt; c.border = border
        ws.row_dimensions[TOTAL_ROW].height = 18

        # ── Anchos de columna ─────────────────────────────────────────────
        ws.column_dimensions[get_column_letter(COL_PRECIO)].width = 12
        ws.column_dimensions[get_column_letter(COL_DESC)].width   = 28
        ws.column_dimensions[get_column_letter(COL_REF)].width    = 10
        for col_j in range(COL_FIRST_STORE, COL_TOTAL):
            ws.column_dimensions[get_column_letter(col_j)].width = 14
        ws.column_dimensions[get_column_letter(COL_TOTAL)].width  = 14

        # Congelar paneles: fijar primeras 3 filas + 3 columnas
        ws.freeze_panes = ws.cell(HDR_ROW + 1, COL_FIRST_STORE)

    # ── Serializar y devolver ─────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    tipo_label_file = "Unidades" if is_units else "Costos"
    filename = f"{collection.name} - {tipo_label_file}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Sales summary (gerencia) ─────────────────────────────────────────────────

@router.get("/summary/by-reference", tags=["reports"])
def sales_by_reference(
    collection_id: Optional[int]  = None,
    vendor_id:     Optional[int]  = None,
    category:      Optional[str]  = None,
    date_from:     Optional[date] = None,
    date_to:       Optional[date] = None,
    db: Session = Depends(get_db),
    _:  User    = Depends(require_manager),
):
    """
    Retorna ventas agrupadas por referencia.
    Filtra opcionalmente por colección, vendedor, categoría y rango de fechas.
    """
    from models.reference import Reference
    q = (
        db.query(
            Reference.code,
            Reference.description,
            Reference.category,
            func.sum(OrderLine.quantity).label("total_units"),
            func.sum(OrderLine.quantity * OrderLine.unit_price).label("total_value"),
        )
        .join(OrderLine, OrderLine.reference_id == Reference.id)
        .join(Order, Order.id == OrderLine.order_id)
        .filter(Order.status != OrderStatus.CANCELLED)
    )
    if collection_id:
        q = q.filter(Order.collection_id == collection_id)
    if vendor_id:
        q = q.filter(Order.vendor_id == vendor_id)
    if category:
        q = q.filter(Reference.category == category)
    if date_from:
        q = q.filter(cast(Order.created_at, Date) >= date_from)
    if date_to:
        q = q.filter(cast(Order.created_at, Date) <= date_to)

    rows = q.group_by(Reference.id).order_by(func.sum(OrderLine.quantity).desc()).all()
    return [
        {
            "code":        r.code,
            "description": r.description,
            "category":    r.category,
            "total_units": int(r.total_units),
            "total_value": float(r.total_value),
        }
        for r in rows
    ]


@router.get("/summary/by-collection", tags=["reports"])
def sales_by_collection(
    db: Session = Depends(get_db),
    _:  User    = Depends(require_manager),
):
    """
    Métricas consolidadas por colección para el panel de comparativas.
    Retorna, por cada colección que tenga al menos un pedido no cancelado:
      - total de pedidos, unidades vendidas, valor total
      - desglose de unidades por categoría de prenda
      - nombre y año/temporada de la colección
    """
    from models.collection import Collection
    from models.reference import Reference

    # ── Totales por colección ────────────────────────────────────────────────
    totals_q = (
        db.query(
            Collection.id,
            Collection.name,
            Collection.year,
            Collection.season,
            func.count(Order.id.distinct()).label("order_count"),
            func.sum(OrderLine.quantity).label("total_units"),
            func.sum(OrderLine.quantity * OrderLine.unit_price).label("total_value"),
        )
        .join(Order, Order.collection_id == Collection.id)
        .join(OrderLine, OrderLine.order_id == Order.id)
        .filter(Order.status != OrderStatus.CANCELLED)
        .group_by(Collection.id)
        .order_by(Collection.year.desc(), Collection.season.desc())
        .all()
    )

    # ── Desglose por categoría (todas las colecciones de una vez) ────────────
    cat_q = (
        db.query(
            Order.collection_id,
            Reference.category,
            func.sum(OrderLine.quantity).label("units"),
            func.sum(OrderLine.quantity * OrderLine.unit_price).label("value"),
        )
        .join(OrderLine, OrderLine.order_id == Order.id)
        .join(Reference, Reference.id == OrderLine.reference_id)
        .filter(Order.status != OrderStatus.CANCELLED)
        .group_by(Order.collection_id, Reference.category)
        .all()
    )

    # Indexar por collection_id
    cat_by_col: dict[int, list] = {}
    for row in cat_q:
        cat_by_col.setdefault(row.collection_id, []).append({
            "category":    row.category.value if hasattr(row.category, "value") else str(row.category),
            "total_units": int(row.units),
            "total_value": float(row.value),
        })

    return [
        {
            "collection_id":   r.id,
            "collection_name": r.name,
            "year":            r.year,
            "season":          r.season,
            "order_count":     int(r.order_count),
            "total_units":     int(r.total_units),
            "total_value":     float(r.total_value),
            "by_category":     sorted(
                cat_by_col.get(r.id, []),
                key=lambda x: x["total_units"],
                reverse=True,
            ),
        }
        for r in totals_q
    ]


@router.get("/summary/by-vendor", tags=["reports"])
def sales_by_vendor(
    collection_id: Optional[int]  = None,
    date_from:     Optional[date] = None,
    date_to:       Optional[date] = None,
    db: Session = Depends(get_db),
    _:  User    = Depends(require_manager),
):
    """Ventas totales por vendedor. Filtra opcionalmente por colección y rango de fechas."""
    q = (
        db.query(
            User.id,
            User.full_name,
            func.count(Order.id).label("order_count"),
            func.sum(OrderLine.quantity).label("total_units"),
            func.sum(OrderLine.quantity * OrderLine.unit_price).label("total_value"),
        )
        .join(Order, Order.vendor_id == User.id)
        .join(OrderLine, OrderLine.order_id == Order.id)
        .filter(Order.status != OrderStatus.CANCELLED)
    )
    if collection_id:
        q = q.filter(Order.collection_id == collection_id)
    if date_from:
        q = q.filter(cast(Order.created_at, Date) >= date_from)
    if date_to:
        q = q.filter(cast(Order.created_at, Date) <= date_to)

    rows = q.group_by(User.id).all()
    return [
        {
            "vendor_id":   r.id,
            "vendor_name": r.full_name,
            "order_count": int(r.order_count),
            "total_units": int(r.total_units),
            "total_value": float(r.total_value),
        }
        for r in rows
    ]
